from __future__ import annotations

import asyncio
import json
import logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramUnauthorizedError
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import CallbackQuery, FSInputFile, Message

from app.config import BIOLOGIYA_PARTS, BLOCKS, BRANCHES, DIRECT_GROUP_SUBJECTS, ENGLISH_BEGINNER_PARTS, ENGLISH_ELEMENTARY_PARTS, ENGLISH_INTERMEDIATE_PARTS, ENGLISH_LEVELS, ENGLISH_PREINTERMEDIATE_PARTS, HAMSHIRALIK_PARTS, IT_PARTS, KAMPYUTER_PARTS, KIMYO_PARTS, LEVELS, RUS_A1_PARTS, RUS_A2_PARTS, RUS_B1_PARTS, RUS_LEVELS, SUBJECTS, get_settings
from app.db import Database
from app.exam_service import QuestionBank, percent_text, question_text, remaining_seconds
from app.keyboards import exam_options_keyboard, main_menu, simple_inline

logging.basicConfig(level=logging.INFO)

settings = get_settings()
db = Database(settings.database_path)
question_bank = QuestionBank(settings.questions_path)
router = Router()

class GroupCreateStates(StatesGroup):
    choosing_branch = State()
    choosing_subject = State()
    entering_group = State()

class ExamStates(StatesGroup):
    choosing_branch = State()
    choosing_subject = State()
    choosing_level = State()
    choosing_block = State()
    choosing_group = State()
    entering_name = State()

class ResultStates(StatesGroup):
    choosing_branch = State()
    choosing_subject = State()
    choosing_group = State()

class ExportStates(StatesGroup):
    choosing_branch = State()

def is_admin(user_id: int) -> bool:
    return user_id in settings.admin_ids

async def send_main_menu(message: Message) -> None:
    await message.answer(
        "Asosiy menyu:",
        reply_markup=main_menu(is_admin(message.from_user.id if message.from_user else 0)),
    )


BRANCH_LABELS = [
    '1️⃣ Niyozbosh',
    '2️⃣ Xalqabod',
    '3️⃣ Gulbahor',
    '4️⃣ Kasblar',
    '5️⃣ Kids1',
    '6️⃣ Kids2',
    '7️⃣ Do’stobod',
    '8️⃣ Olmazor',
    '9️⃣ Chinoz',
    '1️⃣0️⃣ Krasin',
    '1️⃣1️⃣ Pitiletka',
    '1️⃣2️⃣ Qo’rg’oncha',
    '1️⃣3️⃣ Kids 3',
]


def branch_inline_items(prefix: str) -> list[tuple[str, str]]:
    return [(label, f'{prefix}:{idx}') for idx, label in enumerate(BRANCH_LABELS)]


def _display_dt(value: str | None) -> str:
    if not value:
        return ''
    try:
        dt = datetime.fromisoformat(value)
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return value


def _safe_sheet_title(name: str) -> str:
    banned = set('[]:*?/\\')
    cleaned = ''.join('_' if ch in banned else ch for ch in name).strip()
    return cleaned[:31] or 'Sheet'


def build_branch_results_workbook(branch: str) -> Path:
    rows = db.list_export_attempts_for_branch(branch)
    out_dir = settings.questions_path.parent / 'exports'
    out_dir.mkdir(parents=True, exist_ok=True)
    file_path = out_dir / f"results_{branch.replace(' ', '_')}.xlsx"

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    grouped_by_subject: dict[str, list] = defaultdict(list)
    for row in rows:
        grouped_by_subject[row['subject_name']].append(row)

    header_fill = PatternFill('solid', fgColor='D9E2F3')
    first_fill = PatternFill('solid', fgColor='70AD47')
    second_fill = PatternFill('solid', fgColor='5B9BD5')
    third_fill = PatternFill('solid', fgColor='FFC000')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    headers = ['№', 'Ism familya', 'Foiz', 'Gurux', 'Daraja', "Bo'lim", 'Boshlangan', 'Yakunlangan']

    if not grouped_by_subject:
        ws = wb.create_sheet('Natijalar')
        ws['A1'] = f'{branch} filialida natijalar topilmadi.'
        wb.save(file_path)
        return file_path

    for subject_name, subject_rows in grouped_by_subject.items():
        ws = wb.create_sheet(_safe_sheet_title(subject_name))
        current_row = 1
        blocks: dict[tuple[str, str, str], list] = defaultdict(list)
        for row in subject_rows:
            blocks[(row['group_number'], row['level_name'], row['block_name'])].append(row)

        def sort_key(item):
            group_number, level_name, block_name = item[0]
            try:
                gk = (0, int(group_number))
            except Exception:
                gk = (1, str(group_number))
            return (gk, str(level_name).lower(), str(block_name).lower())

        for idx_block, (key, attempts) in enumerate(sorted(blocks.items(), key=sort_key)):
            group_number, level_name, block_name = key
            attempts_sorted = sorted(
                attempts,
                key=lambda r: ((r['score'] / r['total_questions']) if r['total_questions'] else 0, r['submitted_at'] or '', r['id']),
                reverse=True,
            )
            if idx_block > 0:
                current_row += 2
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center
            for pos, attempt in enumerate(attempts_sorted, start=1):
                current_row += 1
                percent = round((attempt['score'] / attempt['total_questions']) * 100) if attempt['total_questions'] else 0
                values = [
                    pos,
                    attempt['full_name'],
                    f'{percent}%',
                    group_number,
                    level_name,
                    block_name,
                    _display_dt(attempt['started_at']),
                    _display_dt(attempt['submitted_at']),
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = left if col_idx == 2 else center
                fill = first_fill if pos == 1 else second_fill if pos == 2 else third_fill if pos == 3 else None
                if fill:
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=current_row, column=col_idx).fill = fill
        widths = {'A': 6, 'B': 32, 'C': 10, 'D': 10, 'E': 16, 'F': 18, 'G': 22, 'H': 22}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    wb.save(file_path)
    return file_path

@router.message(CommandStart())
async def start_handler(message: Message, state: FSMContext) -> None:
    if not message.from_user:
        return
    db.upsert_user(message.from_user.id, message.from_user.full_name, message.from_user.username)
    await state.clear()
    caption = "<b>Oylik imtihonga xush kelibsiz</b>"
    photo_path = settings.questions_path.parent / 'start_banner.jpg'
    if photo_path.exists():
        await message.answer_photo(
            FSInputFile(photo_path),
            caption=caption,
            reply_markup=main_menu(is_admin(message.from_user.id)),
        )
    else:
        await message.answer(caption, reply_markup=main_menu(is_admin(message.from_user.id)))

@router.message(F.text == 'Guruh yaratish')
async def group_create_start(message: Message, state: FSMContext) -> None:
    if not message.from_user or not is_admin(message.from_user.id):
        await message.answer("Bu bo'lim faqat adminlar uchun.")
        return
    await state.set_state(GroupCreateStates.choosing_branch)
    await message.answer(
        'Qaysi filial uchun guruh yaratmoqchisiz?',
        reply_markup=simple_inline(branch_inline_items('gc_branch')),
    )

@router.callback_query(GroupCreateStates.choosing_branch, F.data.startswith('gc_branch:'))
async def gc_branch_selected(callback: CallbackQuery, state: FSMContext) -> None:
    idx = int(callback.data.split(':')[1])
    branch = BRANCHES[idx]
    await state.update_data(branch=branch)
    await state.set_state(GroupCreateStates.choosing_subject)
    await callback.message.edit_text(
        f'Filial: <b>{branch}</b>\n\nQaysi fan uchun guruh yaratmoqchisiz?',
        reply_markup=simple_inline([(name, f'gc_subj:{idx}') for idx, name in enumerate(SUBJECTS)]),
    )
    await callback.answer()

@router.callback_query(GroupCreateStates.choosing_subject, F.data.startswith('gc_subj:'))
async def gc_subject_selected(callback: CallbackQuery, state: FSMContext) -> None:
    idx = int(callback.data.split(':')[1])
    subject_name = SUBJECTS[idx]
    await state.update_data(subject_name=subject_name)
    await state.set_state(GroupCreateStates.entering_group)
    data = await state.get_data()
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\n"
        "Endi guruh raqamlarini yozing.\nMasalan: <code>101,102,103,104,105</code>",
    )
    await callback.answer()

@router.message(GroupCreateStates.entering_group)
async def gc_group_number_entered(message: Message, state: FSMContext) -> None:
    if not message.from_user:
        return
    raw = (message.text or '').strip()
    if not raw:
        await message.answer("Guruh raqamini yozing.")
        return

    group_numbers = [chunk.strip() for chunk in raw.replace('\n', ',').split(',') if chunk.strip()]
    if not group_numbers:
        await message.answer("Kamida bitta guruh raqamini yozing.")
        return

    data = await state.get_data()
    created: list[str] = []
    duplicates: list[str] = []
    for group_number in group_numbers:
        ok, _ = db.create_group(data['branch'], data['subject_name'], group_number, message.from_user.id)
        if ok:
            created.append(group_number)
        else:
            duplicates.append(group_number)

    await state.clear()
    lines = [f"🏫 Filial: {data['branch']}", f"📚 Fan: {data['subject_name']}"]
    if created:
        lines.append('✅ Saqlandi: ' + ', '.join(created))
    if duplicates:
        lines.append('⚠️ Avvaldan bor: ' + ', '.join(duplicates))
    await message.answer('\n'.join(lines))
    await send_main_menu(message)



@router.message(F.text == 'Excel format yuklash')
async def export_start(message: Message, state: FSMContext) -> None:
    if not message.from_user or not is_admin(message.from_user.id):
        await message.answer("Bu bo'lim faqat adminlar uchun.")
        return
    branches = BRANCHES
    await state.set_state(ExportStates.choosing_branch)
    await message.answer(
        'Qaysi filial natijalarini yuklamoqchisiz?',
        reply_markup=simple_inline(branch_inline_items('exp_branch')),
    )

@router.callback_query(ExportStates.choosing_branch, F.data.startswith('exp_branch:'))
async def export_branch_selected(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    branches = BRANCHES
    idx = int(callback.data.split(':')[1])
    if idx >= len(branches):
        await state.clear()
        await callback.message.edit_text('Filial topilmadi.')
        await callback.answer()
        return
    branch = branches[idx]
    await callback.message.edit_text('Excel tayyorlanmoqda...')
    path = build_branch_results_workbook(branch)
    await state.clear()
    await callback.message.answer_document(FSInputFile(path), caption=f'{branch} filial natijalari')
    await callback.answer()

@router.message(F.text == 'Oylik imtihon')
async def exam_start(message: Message, state: FSMContext) -> None:
    if not message.from_user:
        return
    active = db.find_active_attempt(message.from_user.id)
    if active:
        await message.answer("Sizda tugallanmagan test bor. O'shani davom ettirasiz.")
        await show_question(message.bot, message.from_user.id, active['id'])
        return
    await state.set_state(ExamStates.choosing_branch)
    await message.answer(
        'Filial tanlang:',
        reply_markup=simple_inline(branch_inline_items('ex_branch')),
    )

@router.callback_query(ExamStates.choosing_branch, F.data.startswith('ex_branch:'))
async def ex_branch_selected(callback: CallbackQuery, state: FSMContext) -> None:
    branch = BRANCHES[int(callback.data.split(':')[1])]
    await state.update_data(branch=branch)
    await state.set_state(ExamStates.choosing_subject)
    await callback.message.edit_text(
        f'Filial: <b>{branch}</b>\n\nFan tanlang:',
        reply_markup=simple_inline([(name, f'ex_subj:{idx}') for idx, name in enumerate(SUBJECTS)]),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_subject, F.data.startswith('ex_subj:'))
async def ex_subject_selected(callback: CallbackQuery, state: FSMContext) -> None:
    subject_name = SUBJECTS[int(callback.data.split(':')[1])]
    await state.update_data(subject_name=subject_name)
    await state.set_state(ExamStates.choosing_level)
    data = await state.get_data()

    if subject_name == 'Kimyo':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_kimyo_part:{idx}') for idx, name in enumerate(KIMYO_PARTS)]),
        )
        await callback.answer()
        return

    if subject_name == 'Biologiya':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_bio_part:{idx}') for idx, name in enumerate(BIOLOGIYA_PARTS)]),
        )
        await callback.answer()
        return

    if subject_name == 'Rus tili':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\nDaraja tanlang:",
            reply_markup=simple_inline([(name, f'ex_rus_level:{idx}') for idx, name in enumerate(RUS_LEVELS)]),
        )
        await callback.answer()
        return

    if subject_name == 'Hamshiralik':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_ham_part:{idx}') for idx, name in enumerate(HAMSHIRALIK_PARTS)]),
        )
        await callback.answer()
        return

    if subject_name == 'IT':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_it_part:{idx}') for idx, name in enumerate(IT_PARTS)]),
        )
        await callback.answer()
        return

    if subject_name == 'Kampyuter':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_kamp_part:{idx}') for idx, name in enumerate(KAMPYUTER_PARTS)]),
        )
        await callback.answer()
        return

    if subject_name == 'English':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\nDaraja tanlang:",
            reply_markup=simple_inline([(name, f'ex_eng_level:{idx}') for idx, name in enumerate(ENGLISH_LEVELS)]),
        )
        await callback.answer()
        return

    if subject_name in DIRECT_GROUP_SUBJECTS:
        await state.update_data(level_name='Asosiy', block_name='Asosiy')
        groups = db.list_groups(data['branch'], subject_name)
        if not groups:
            await state.clear()
            await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
            await callback.answer()
            return
        await state.set_state(ExamStates.choosing_group)
        items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\nGuruh tanlang:",
            reply_markup=simple_inline(items),
        )
        await callback.answer()
        return

    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{subject_name}</b>\n\nDaraja tanlang:",
        reply_markup=simple_inline([(name, f'ex_level:{idx}') for idx, name in enumerate(LEVELS)]),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_level, F.data.startswith('ex_rus_level:'))
async def ex_rus_level_selected(callback: CallbackQuery, state: FSMContext) -> None:
    level_name = RUS_LEVELS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=level_name)
    data = await state.get_data()

    if level_name == 'A1':
        await state.set_state(ExamStates.choosing_block)
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nDaraja: <b>{level_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_rus_a1_part:{idx}') for idx, name in enumerate(RUS_A1_PARTS)]),
        )
    elif level_name == 'A2':
        await state.set_state(ExamStates.choosing_block)
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nDaraja: <b>{level_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_rus_a2_part:{idx}') for idx, name in enumerate(RUS_A2_PARTS)]),
        )
    elif level_name == 'B1':
        await state.set_state(ExamStates.choosing_block)
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nDaraja: <b>{level_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_rus_b1_part:{idx}') for idx, name in enumerate(RUS_B1_PARTS)]),
        )
    else:
        await state.clear()
        await callback.message.edit_text(f"<b>{level_name}</b> bo'limi uchun testlar hali kiritilmagan.")
    await callback.answer()

@router.callback_query(ExamStates.choosing_level, F.data.startswith('ex_eng_level:'))
async def ex_english_level_selected(callback: CallbackQuery, state: FSMContext) -> None:
    level_name = ENGLISH_LEVELS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=level_name)
    await state.set_state(ExamStates.choosing_block)
    data = await state.get_data()

    if level_name == 'Starter':
        await state.update_data(block_name='Asosiy')
        groups = db.list_groups(data['branch'], data['subject_name'])
        if not groups:
            await state.clear()
            await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
            await callback.answer()
            return
        await state.set_state(ExamStates.choosing_group)
        items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{level_name}</b>\n\nGuruh tanlang:",
            reply_markup=simple_inline(items),
        )
        await callback.answer()
        return

    if level_name == 'Beginner':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nDaraja: <b>{level_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_eng_beg_part:{idx}') for idx, name in enumerate(ENGLISH_BEGINNER_PARTS)]),
        )
        await callback.answer()
        return

    if level_name == 'Elementary':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nDaraja: <b>{level_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_eng_elem_part:{idx}') for idx, name in enumerate(ENGLISH_ELEMENTARY_PARTS)]),
        )
        await callback.answer()
        return

    if level_name == 'Pre-Intermediate':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nDaraja: <b>{level_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_eng_pre_part:{idx}') for idx, name in enumerate(ENGLISH_PREINTERMEDIATE_PARTS)]),
        )
        await callback.answer()
        return

    if level_name == 'Intermediate':
        await callback.message.edit_text(
            f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nDaraja: <b>{level_name}</b>\n\nBo'lim tanlang:",
            reply_markup=simple_inline([(name, f'ex_eng_int_part:{idx}') for idx, name in enumerate(ENGLISH_INTERMEDIATE_PARTS)]),
        )
        await callback.answer()
        return

    await callback.answer()

@router.callback_query(ExamStates.choosing_level, F.data.startswith('ex_level:'))
async def ex_level_selected(callback: CallbackQuery, state: FSMContext) -> None:
    level_name = LEVELS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=level_name)
    await state.set_state(ExamStates.choosing_block)
    data = await state.get_data()

    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nDaraja: <b>{level_name}</b>\n\nBlok tanlang:",
        reply_markup=simple_inline([(name, f'ex_block:{idx}') for idx, name in enumerate(BLOCKS)]),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_level, F.data.startswith('ex_kimyo_part:'))
async def ex_kimyo_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = KIMYO_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_level, F.data.startswith('ex_bio_part:'))
async def ex_bio_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = BIOLOGIYA_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_block, F.data.startswith('ex_rus_a1_part:'))
async def ex_rus_a1_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = RUS_A1_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_block, F.data.startswith('ex_rus_a2_part:'))
async def ex_rus_a2_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = RUS_A2_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_block, F.data.startswith('ex_rus_b1_part:'))
async def ex_rus_b1_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = RUS_B1_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_block, F.data.startswith('ex_eng_beg_part:'))
async def ex_english_beginner_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = ENGLISH_BEGINNER_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_block, F.data.startswith('ex_eng_elem_part:'))
async def ex_english_elementary_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = ENGLISH_ELEMENTARY_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_level, F.data.startswith('ex_ham_part:'))
async def ex_hamshiralik_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = HAMSHIRALIK_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_level, F.data.startswith('ex_it_part:'))
async def ex_it_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = IT_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_block, F.data.startswith('ex_eng_pre_part:'))
async def ex_english_preintermediate_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = ENGLISH_PREINTERMEDIATE_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_block, F.data.startswith('ex_eng_int_part:'))
async def ex_english_intermediate_part_selected(callback: CallbackQuery, state: FSMContext) -> None:
    part_name = ENGLISH_INTERMEDIATE_PARTS[int(callback.data.split(':')[1])]
    await state.update_data(level_name=part_name, block_name='Asosiy')
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nBo'lim: <b>{part_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_block, F.data.startswith('ex_block:'))
async def ex_block_selected(callback: CallbackQuery, state: FSMContext) -> None:
    block_name = BLOCKS[int(callback.data.split(':')[1])]
    await state.update_data(block_name=block_name)
    data = await state.get_data()
    groups = db.list_groups(data['branch'], data['subject_name'])
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu filial va fan uchun hali admin guruh yaratmagan.")
        await callback.answer()
        return
    await state.set_state(ExamStates.choosing_group)
    items = [(row['group_number'], f'ex_group:{row["id"]}') for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{data['branch']}</b>\nFan: <b>{data['subject_name']}</b>\nDaraja: <b>{data['level_name']}</b>\nBlok: <b>{block_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ExamStates.choosing_group, F.data.startswith('ex_group:'))
async def ex_group_selected(callback: CallbackQuery, state: FSMContext) -> None:
    group_id = int(callback.data.split(':')[1])
    group = db.get_group(group_id)
    if not group:
        await callback.answer('Guruh topilmadi.', show_alert=True)
        return
    data = await state.get_data()
    questions = question_bank.get_questions(data['subject_name'], data['level_name'], data['block_name'])
    if not questions:
        await state.clear()
        await callback.message.edit_text(
            "Bu fan/daraja/blok uchun savollar topilmadi. data/questions.json ichiga savollarni kiriting."
        )
        await callback.answer()
        return
    await state.update_data(group_id=group_id, group_number=group['group_number'])
    user_row = db.get_user(callback.from_user.id)
    saved_name = user_row['full_name'] if user_row and user_row['full_name'] else None
    await state.set_state(ExamStates.entering_name)
    if saved_name:
        await callback.message.edit_text(
            f"Sizning saqlangan ism-familyangiz: <b>{saved_name}</b>\n\n"
            "Ism familyangizni to'liq kiriting yoki yangilang."
        )
    else:
        await callback.message.edit_text("Ism familyangizni to'liq kiriting:")
    await callback.answer()

@router.message(ExamStates.entering_name)
async def ex_name_entered(message: Message, state: FSMContext) -> None:
    if not message.from_user:
        return
    data = await state.get_data()
    user_row = db.get_user(message.from_user.id)
    text = (message.text or '').strip()
    if len(text) < 3:
        await message.answer("Ism familyangizni to'liqroq kiriting.")
        return
    full_name = text
    questions = question_bank.get_questions(data['subject_name'], data['level_name'], data['block_name'])
    attempt_id = db.create_attempt(
        user_id=message.from_user.id,
        username=message.from_user.username,
        full_name=full_name,
        branch=data['branch'],
        subject_name=data['subject_name'],
        level_name=data['level_name'],
        block_name=data['block_name'],
        group_id=data['group_id'],
        group_number=data['group_number'],
        questions=questions,
    )
    db.upsert_user(message.from_user.id, full_name, message.from_user.username)
    await state.clear()
    await message.answer(
        f"✅ Test boshlandi. Sizda {settings.exam_minutes} minut vaqt bor.\n"
        f"Savollar soni: {len(questions)}\n"
        "Vaqt tugasa test avtomatik yakunlanadi."
    )
    await show_question(message.bot, message.from_user.id, attempt_id)

async def show_question(bot: Bot, user_id: int, attempt_id: int) -> None:
    attempt = db.get_attempt(attempt_id)
    if not attempt:
        await bot.send_message(user_id, 'Test topilmadi.')
        return
    if attempt['status'] != 'active':
        await bot.send_message(user_id, "✅ Testingiz muvaffaqiyatli qabul qilindi. Balingizni Natijalar bo'limidan ko'ra olasiz.")
        return
    if remaining_seconds(attempt['started_at'], settings.exam_minutes) <= 0:
        await finish_attempt(bot, user_id, attempt_id, by_time=True)
        return
    question_index = db.get_current_question_index(attempt_id)
    questions = json.loads(attempt['questions_json'])
    q = questions[question_index]
    text = question_text(dict(attempt), question_index, settings.exam_minutes)
    await bot.send_message(
        user_id,
        text,
        reply_markup=exam_options_keyboard(attempt_id, question_index, q['options']),
    )

@router.callback_query(F.data.startswith('ea:'))
async def exam_answer(callback: CallbackQuery) -> None:
    _, attempt_id_raw, question_index_raw, option_index_raw = callback.data.split(':')
    attempt_id = int(attempt_id_raw)
    question_index = int(question_index_raw)
    option_index = int(option_index_raw)
    attempt = db.get_attempt(attempt_id)
    if not attempt or attempt['user_id'] != callback.from_user.id:
        await callback.answer('Bu test sizga tegishli emas.', show_alert=True)
        return
    if attempt['status'] != 'active':
        await callback.answer('Test allaqachon tugagan.', show_alert=True)
        return
    if remaining_seconds(attempt['started_at'], settings.exam_minutes) <= 0:
        await finish_attempt(callback.bot, callback.from_user.id, attempt_id, by_time=True)
        await callback.answer()
        return
    db.save_answer(attempt_id, question_index, option_index)
    answered_count = db.get_answered_count(attempt_id)
    if answered_count >= attempt['total_questions']:
        await finish_attempt(callback.bot, callback.from_user.id, attempt_id)
        await callback.answer('Test yakunlandi.')
        return
    questions = json.loads(attempt['questions_json'])
    next_question = min(question_index + 1, len(questions) - 1)
    text = question_text(dict(db.get_attempt(attempt_id)), next_question, settings.exam_minutes)
    await callback.message.edit_text(
        text,
        reply_markup=exam_options_keyboard(attempt_id, next_question, questions[next_question]['options']),
    )
    await callback.answer('Javob saqlandi.')

async def finish_attempt(bot: Bot, user_id: int, attempt_id: int, by_time: bool = False) -> None:
    status = 'time_up' if by_time else 'submitted'
    attempt = db.finalize_attempt(attempt_id, status=status)
    if not attempt:
        await bot.send_message(user_id, 'Testni yakunlashda xatolik yuz berdi.')
        return
    if by_time:
        await bot.send_message(user_id, "⏰ Vaqt tugadi. Testingiz avtomatik qabul qilindi. Balingizni Natijalar bo'limidan ko'ra olasiz.")
    else:
        await bot.send_message(user_id, "✅ Testingiz muvaffaqiyatli qabul qilindi. Balingizni Natijalar bo'limidan ko'ra olasiz.")

@router.message(F.text == 'Natijalar')
async def results_start(message: Message, state: FSMContext) -> None:
    await state.set_state(ResultStates.choosing_branch)
    await message.answer(
        'Filial tanlang:',
        reply_markup=simple_inline(branch_inline_items('res_branch')),
    )

@router.callback_query(ResultStates.choosing_branch, F.data.startswith('res_branch:'))
async def res_branch(callback: CallbackQuery, state: FSMContext) -> None:
    branch = BRANCHES[int(callback.data.split(':')[1])]
    await state.update_data(branch=branch)
    subjects = db.list_branch_subjects(branch)
    if not subjects:
        await state.clear()
        await callback.message.edit_text("Bu filial uchun hali fanlar topilmadi.")
        await callback.answer()
        return
    await state.set_state(ResultStates.choosing_subject)
    await callback.message.edit_text(
        f"Filial: <b>{branch}</b>\n\nFan tanlang:",
        reply_markup=simple_inline([(name, f'res_subject:{idx}') for idx, name in enumerate(subjects)]),
    )
    await callback.answer()

@router.callback_query(ResultStates.choosing_subject, F.data.startswith('res_subject:'))
async def res_subject(callback: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    branch = data.get('branch')
    if not branch:
        await state.clear()
        await callback.message.edit_text("Qaytadan /start bosing.")
        await callback.answer()
        return
    subjects = db.list_branch_subjects(branch)
    subject_name = subjects[int(callback.data.split(':')[1])]
    await state.update_data(subject_name=subject_name)
    groups = db.list_groups_for_branch_subject(branch, subject_name)
    if not groups:
        await state.clear()
        await callback.message.edit_text("Bu fan uchun hali guruhlar topilmadi.")
        await callback.answer()
        return
    await state.set_state(ResultStates.choosing_group)
    items = [(row['group_number'], f"res_group:{row['id']}") for row in groups]
    await callback.message.edit_text(
        f"Filial: <b>{branch}</b>\nFan: <b>{subject_name}</b>\n\nGuruh tanlang:",
        reply_markup=simple_inline(items),
    )
    await callback.answer()

@router.callback_query(ResultStates.choosing_group, F.data.startswith('res_group:'))
async def res_group(callback: CallbackQuery, state: FSMContext) -> None:
    group_id = int(callback.data.split(':')[1])
    group = db.get_group(group_id)
    rows = db.list_students_with_results(group_id)
    await state.clear()
    if not group:
        await callback.message.edit_text("Guruh topilmadi.")
        await callback.answer()
        return
    if not rows:
        await callback.message.edit_text("Bu guruh bo'yicha hali natijalar yo'q.")
        await callback.answer()
        return

    top_three = rows[:3]
    other_rows = rows[3:]
    lines = [
        '🏆 <b>Guruh natijalari</b>',
        f"🏫 Filial: {group['branch']}",
        f"📚 Fan: {group['subject_name']}",
        f"👥 Guruh: {group['group_number']}",
        '',
    ]
    medals = ["🥇 1-o'rin", "🥈 2-o'rin", "🥉 3-o'rin"]
    for idx, row in enumerate(top_three):
        lines.append(f"{medals[idx]} — {row['full_name']} | <b>{percent_text(row['score'], row['total_questions'])}</b>")
    if other_rows:
        lines.append('')
        lines.append("<b>Barcha o'quvchilar:</b>")
        for pos, row in enumerate(other_rows, start=4):
            lines.append(f"{pos}. {row['full_name']} — <b>{percent_text(row['score'], row['total_questions'])}</b>")

    await callback.message.edit_text('\n'.join(lines))
    await callback.answer()

@router.message()
async def fallback_handler(message: Message) -> None:
    await message.answer(
        "Buyruqni menyudan tanlang yoki /start bosing.",
        reply_markup=main_menu(is_admin(message.from_user.id if message.from_user else 0)),
    )

async def expiration_watcher(bot: Bot) -> None:
    while True:
        expired = db.get_expired_attempts(settings.exam_minutes)
        for attempt in expired:
            await finish_attempt(bot, attempt['user_id'], attempt['id'], by_time=True)
        await asyncio.sleep(10)

async def main() -> None:
    db.init()
    bot = Bot(token=settings.bot_token, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)

    async def _startup() -> None:
        asyncio.create_task(expiration_watcher(bot))

    try:
        await bot.delete_webhook(drop_pending_updates=False)
        await _startup()
        await dp.start_polling(bot)
    except TelegramUnauthorizedError:
        print()
        print('=' * 72)
        print("XATOLIK: Telegram bot token noto'g'ri yoki bekor qilingan.")
        print()
        print('1) @BotFather dan yangi token oling')
        print('2) monthly_exam_bot papkasidagi .env faylni oching')
        print('3) BOT_TOKEN=... qatoriga yangi tokenni yozing')
        print('4) Keyin botni qayta ishga tushiring')
        print('=' * 72)
        raise SystemExit(1)

if __name__ == '__main__':
    asyncio.run(main())
